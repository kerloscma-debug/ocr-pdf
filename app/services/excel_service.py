"""Generate Excel with 2 sheets: detail + summary."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COLS = [
    ("Batch ID",14),("صفحة",6),("التصنيف",18),("المورد",30),
    ("الرقم الضريبي",16),("الاسم التجاري",24),("رقم الفاتورة",20),
    ("تاريخ الفاتورة",14),("قبل الضريبة",14),("الضريبة",12),
    ("الإجمالي",12),("QR موجود",10),("QR محلل",10),("QR متسق",10),
    ("ثقة",8),("VAT صالح",10),("حساب صحيح",10),("نسبة 15%",10),
    ("مراجعة يدوية",12),("ملاحظات",30),
]

thin = Side(style="thin",color="AAAAAA")
BDR  = Border(left=thin,right=thin,top=thin,bottom=thin)

def _h(ws,row,col,val,width=None):
    c=ws.cell(row=row,column=col,value=val)
    c.font=Font(bold=True,color="FFFFFF",name="Arial",size=9)
    c.fill=PatternFill("solid",fgColor="1F3864")
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    c.border=BDR
    if width: ws.column_dimensions[get_column_letter(col)].width=width

def _d(ws,row,col,val,fmt=None,bg=None):
    c=ws.cell(row=row,column=col,value=val)
    c.font=Font(name="Arial",size=9)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    c.border=BDR
    if fmt: c.number_format=fmt
    if bg: c.fill=PatternFill("solid",fgColor=bg)
    return c

def generate(invoices: list, batch_id: str, out_path: Path):
    wb=Workbook()
    ws=wb.active; ws.title="Invoices"; ws.sheet_view.rightToLeft=True
    for i,(label,width) in enumerate(COLS,1):
        _h(ws,1,i,label,width)
    ws.row_dimensions[1].height=32

    summary_data={}
    for idx,inv in enumerate(invoices,2):
        needs=getattr(inv,"requires_manual_review",False)
        bg="FFE7E7" if needs else ("FFF2CC" if getattr(inv,"is_duplicate",False) else None)
        row_vals=[
            batch_id, inv.page_number, inv.classification, inv.supplier_name,
            inv.vat_number, inv.trade_name, inv.invoice_number, inv.invoice_date,
            inv.amount_before_vat, inv.vat_amount, inv.total_amount,
            "نعم" if inv.qr_present else "لا",
            "نعم" if inv.qr_parsed else "لا",
            "نعم" if inv.qr_consistent else ("لا" if inv.qr_consistent is False else "—"),
            inv.confidence,
            "✓" if inv.vat_number_valid else "✗",
            "✓" if inv.math_valid else "✗",
            "✓" if inv.vat_rate_valid else "✗",
            "نعم" if needs else "لا",
            inv.notes,
        ]
        for c,v in enumerate(row_vals,1):
            fmt="#,##0.00" if c in (9,10,11) else None
            _d(ws,idx,c,v,fmt=fmt,bg=bg)
        # summary
        key=(inv.vat_number or "—", inv.trade_name or "—")
        if key not in summary_data:
            summary_data[key]={"count":0,"base":0,"vat":0,"total":0}
        summary_data[key]["count"]+=1
        summary_data[key]["base"]+=(inv.amount_before_vat or 0)
        summary_data[key]["vat"]+=(inv.vat_amount or 0)
        summary_data[key]["total"]+=(inv.total_amount or 0)

    ws.auto_filter.ref=f"A1:{get_column_letter(len(COLS))}1"
    ws.freeze_panes="A2"

    # Summary sheet
    ws2=wb.create_sheet("Summary"); ws2.sheet_view.rightToLeft=True
    s_hdrs=[("الرقم الضريبي",18),("الاسم التجاري",28),("عدد الفواتير",14),
            ("إجمالي قبل الضريبة",18),("إجمالي الضريبة",16),("الإجمالي الكلي",16)]
    for i,(lbl,w) in enumerate(s_hdrs,1):
        _h(ws2,1,i,lbl,w)
    for r,(key,vals) in enumerate(summary_data.items(),2):
        for c,v in enumerate([key[0],key[1],vals["count"],
                               vals["base"],vals["vat"],vals["total"]],1):
            fmt="#,##0.00" if c>3 else None
            _d(ws2,r,c,v,fmt=fmt)
    ws2.freeze_panes="A2"

    out_path.parent.mkdir(exist_ok=True)
    wb.save(str(out_path))
